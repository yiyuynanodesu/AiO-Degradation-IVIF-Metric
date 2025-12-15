import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill

"""
合并所有去退化IVIF任务detail文件（就是每个退化类别） 仅适用于去退化IVIF
"""

def merge_excel_files(file_paths, output_file):
    """
    合并多个Excel文件，提取每个sheet的平均值，并按要求格式化输出
    
    Args:
        file_paths: Excel文件路径列表
        output_file: 输出文件路径
    """
    
    # 创建一个新的工作簿
    wb = Workbook()
    
    # 存储所有文件的数据
    all_data = {}  # {sheet_name: {file_name: [EI, EN, SF, AG, SD, CLIPIQA]}}
    
    # 遍历所有输入文件
    for file_path in file_paths:
        file_name = os.path.basename(file_path)
        
        try:
            # 读取Excel文件的所有sheet
            xls = pd.ExcelFile(file_path)
            
            # 遍历每个sheet
            for sheet_name in xls.sheet_names:
                # 跳过名为'Sheet'的空sheet
                if sheet_name == 'Sheet':
                    # 尝试读取sheet，如果失败或为空则跳过
                    try:
                        df = pd.read_excel(xls, sheet_name=sheet_name)
                        if df.empty:
                            continue
                    except:
                        continue
                
                # 读取sheet数据
                df = pd.read_excel(xls, sheet_name=sheet_name)
                
                # 检查数据是否为空
                if df.empty:
                    continue
                
                # 确保列数足够
                if df.shape[1] >= 7:  # 至少有7列
                    # 获取最后一行（平均值行）
                    last_row = df.iloc[-1]
                    
                    # 提取指标值
                    # 判断最后一行第一列是否是平均值标识
                    first_cell_value = str(last_row.iloc[0]).lower() if pd.notna(last_row.iloc[0]) else ""
                    
                    if any(keyword in first_cell_value for keyword in ['mean', 'average', 'avg', '平均值', '均值']):
                        # 从第二列开始提取指标值
                        values = []
                        for i in range(1, 7):  # 提取第2-7列
                            if i < len(last_row):
                                values.append(last_row.iloc[i])
                            else:
                                values.append(None)
                    else:
                        # 可能最后一行就是数值，直接提取前5列
                        values = []
                        for i in range(0, 6):  # 提取第1-6列
                            if i < len(last_row):
                                values.append(last_row.iloc[i])
                            else:
                                values.append(None)
                    
                    # 确保有5个值且都是数值
                    valid_values = []
                    for val in values:
                        if pd.isna(val):
                            valid_values.append(None)
                        elif isinstance(val, (int, float)):
                            valid_values.append(float(val))
                        else:
                            # 尝试转换为数值
                            try:
                                valid_values.append(float(val))
                            except:
                                valid_values.append(None)
                    
                    # 检查是否有足够有效的数值
                    if len([v for v in valid_values if v is not None]) >= 3:  # 至少3个有效值
                        # 初始化数据结构
                        if sheet_name not in all_data:
                            all_data[sheet_name] = {}
                        
                        # 存储数据
                        all_data[sheet_name][file_name] = valid_values
        except Exception as e:
            print(f"处理文件 {file_name} 时出错: {e}")
    
    # 如果没有数据，则返回
    if not all_data:
        print("没有找到有效数据")
        return
    
    # 移除默认的sheet（如果有）
    if "Sheet" in wb.sheetnames:
        default_ws = wb["Sheet"]
        wb.remove(default_ws)
    
    # 为每个sheet创建数据
    for sheet_name, file_data in all_data.items():
        # 创建新sheet
        ws = wb.create_sheet(title=sheet_name)
        
        # 写入表头
        headers = ["文件名", "EI", "EN", "SF", "AG", "SD", "CLIPIQA"]
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
        
        # 写入数据
        row_idx = 2
        for file_name, values in file_data.items():
            ws.cell(row=row_idx, column=1, value=file_name)
            for col_idx, value in enumerate(values, 2):
                ws.cell(row=row_idx, column=col_idx, value=value)
            row_idx += 1
    
    # 为每个sheet设置颜色标记
    for sheet_name in all_data.keys():
        ws = wb[sheet_name]
        
        # 获取数据范围
        max_row = ws.max_row
        if max_row <= 1:  # 只有表头
            continue
        
        # 为每个指标列（B到G列）找出前三大值并设置颜色
        for col_idx in range(2, 8):  # B=2, C=3, D=4, E=5, F=6, G=7
            # 获取该列的所有值（排除表头）
            values = []
            for row in range(2, max_row + 1):
                cell_value = ws.cell(row=row, column=col_idx).value
                if cell_value is not None:
                    try:
                        # 确保是数值类型
                        float_val = float(cell_value)
                        values.append((row, float_val))
                    except (ValueError, TypeError):
                        continue
            
            if len(values) < 3:  # 至少需要3个值才能标记前三大
                continue
            
            # 按值排序（降序）
            sorted_values = sorted(values, key=lambda x: x[1], reverse=True)
            
            # 定义颜色填充
            red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")  # 红色
            blue_fill = PatternFill(start_color="FF0000FF", end_color="FF0000FF", fill_type="solid")  # 蓝色
            green_fill = PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid")  # 绿色
            
            # 应用颜色
            # 最大值（红色）
            max_row_idx = sorted_values[0][0]
            ws.cell(row=max_row_idx, column=col_idx).fill = red_fill
            
            # 找出第二大值（排除与最大值相同的值）
            second_max_value = None
            second_max_row_idx = None
            for row_idx, value in sorted_values[1:]:
                if value != sorted_values[0][1]:
                    second_max_value = value
                    second_max_row_idx = row_idx
                    break
            
            # 第二大值（蓝色）
            if second_max_row_idx is not None:
                ws.cell(row=second_max_row_idx, column=col_idx).fill = blue_fill
                
                # 找出第三大值（排除与前两个相同的值）
                third_max_value = None
                third_max_row_idx = None
                for row_idx, value in sorted_values[2:]:
                    if value != sorted_values[0][1] and value != second_max_value:
                        third_max_value = value
                        third_max_row_idx = row_idx
                        break
                
                # 第三大值（绿色）
                if third_max_row_idx is not None:
                    ws.cell(row=third_max_row_idx, column=col_idx).fill = green_fill
    
    # 保存工作簿
    wb.save(output_file)
    print(f"合并完成，结果已保存到: {output_file}")
    
    # 显示统计信息
    print(f"\n处理统计:")
    for sheet_name in all_data.keys():
        print(f"  {sheet_name}: {len(all_data[sheet_name])} 个文件")

# 使用示例
if __name__ == "__main__":
    # 获取要合并的路径
    target_path = input('请输入要合并的路径：').strip()
    
    # 如果路径为空，使用当前目录
    if not target_path:
        target_path = "."
    
    # 获取所有Excel文件
    excel_files = []
    for file in os.listdir(target_path):
        if file.lower().endswith(('.xlsx', '.xls')):
            excel_files.append(os.path.join(target_path, file))
    
    if excel_files:
        print(f"找到 {len(excel_files)} 个Excel文件:")
        for file in excel_files:
            print(f"  - {os.path.basename(file)}")
        
        # 输出文件路径
        output_path = "detail_merge_result.xlsx"
        
        # 执行合并
        merge_excel_files(excel_files, output_path)
        
        print(f"\n处理完成!")
        print(f"输出文件: {output_path}")
    else:
        print(f"在路径 '{target_path}' 中未找到Excel文件")