import pandas as pd
import numpy as np
import os
import warnings
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl

warnings.filterwarnings('ignore')

def highlight_max_values(ws, df, sheet_name):
    """
    在工作表中高亮显示最大值和第二大值
    """
    # 定义填充颜色
    max_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 黄色
    second_max_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # 绿色
    
    # 对于每一列（排除第一列模型名）
    for col_idx in range(2, df.shape[1] + 2):  # 从第2列开始（第1列是模型名）
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        col_name = df.columns[col_idx-2]  # 获取列名
        
        # 收集该列的值（跳过标题行）
        values = []
        for row_idx in range(2, ws.max_row + 1):  # 从第2行开始（第1行是标题）
            cell_value = ws[f"{col_letter}{row_idx}"].value
            try:
                # 尝试转换为数值
                if cell_value is not None:
                    values.append(float(cell_value))
                else:
                    values.append(-np.inf)  # 将None视为最小值
            except:
                values.append(-np.inf)  # 无法转换则视为最小值
        
        if not values:
            continue
        
        # 找到最大值
        max_val = max(values)
        max_indices = [i for i, v in enumerate(values) if v == max_val]
        
        # 找到第二大值（排除最大值）
        other_vals = [v for v in values if v < max_val]
        if other_vals:
            second_max_val = max(other_vals)
            second_max_indices = [i for i, v in enumerate(values) if v == second_max_val]
        else:
            # 如果所有值都相等，则没有第二大值
            second_max_val = max_val
            second_max_indices = []
        
        # 高亮最大值
        for idx in max_indices:
            row_idx = idx + 2  # 加2因为行索引从2开始
            cell = ws[f"{col_letter}{row_idx}"]
            cell.fill = max_fill
            cell.font = Font(bold=True)
        
        # 高亮第二大值（如果与最大值不同）
        if second_max_val < max_val:
            for idx in second_max_indices:
                row_idx = idx + 2  # 加2因为行索引从2开始
                cell = ws[f"{col_letter}{row_idx}"]
                cell.fill = second_max_fill
                cell.font = Font(bold=True)

def merge_excel_files(input_folder, output_file):
    """
    合并多个Excel文件，每个文件包含多个sheet，每个sheet表示一个指标
    """
    
    # 存储所有数据
    all_data = {}
    
    # 获取所有Excel文件
    excel_files = [f for f in os.listdir(input_folder) if f.endswith(('.xlsx', '.xls'))]
    
    if not excel_files:
        print(f"在文件夹 {input_folder} 中未找到Excel文件")
        return
    
    print(f"找到 {len(excel_files)} 个Excel文件")
    
    # 处理每个Excel文件
    for file_idx, file_name in enumerate(excel_files, 1):
        file_path = os.path.join(input_folder, file_name)
        model_name = os.path.splitext(file_name)[0]  # 使用文件名作为模型名
        
        print(f"正在处理文件 {file_idx}/{len(excel_files)}: {file_name}")
        
        try:
            # 读取Excel文件的所有sheet
            excel_data = pd.read_excel(file_path, sheet_name=None)

            # 处理每个sheet
            for sheet_name, df in excel_data.items():
  
                # 跳过名为"Sheet"的空sheet
                if sheet_name.lower() == "sheet" and (df.empty or df.shape[0] < 1):
                    continue

                # 跳过空sheet
                if df.empty:
                    continue
                
                # 重置索引
                df = df.reset_index(drop=True)
                
                # 判断数据格式：如果df.shape[0] == 1，说明只有一行数据
                # pandas默认将第一行作为列名，所以实际数据只有一行
                
                # 对于只有一行数据的情况，我们直接使用当前列名
                if df.shape[0] >= 1:
                    # 检查第一行是否都是NaN（可能是空行）
                    if df.iloc[0].isnull().all():
                        # 如果第一行是空行，尝试使用第二行作为数据
                        if df.shape[0] > 1:
                            # 第一行是空行，第二行是数据
                            df = df.iloc[1:].reset_index(drop=True)
                        else:
                            # 只有空行，跳过
                            continue
                    
                    # 此时df应该至少有一行数据
                    if df.empty:
                        continue
                        
                    # 确保列名是字符串类型
                    df.columns = df.columns.astype(str)
                    
                    # 检查第一行是否包含非数值数据（可能是标题行）
                    # 如果第一行的值看起来像标题（包含非数字字符），可能需要特殊处理
                    first_row_str = df.iloc[0].astype(str).str.lower().tolist()
                    has_strings = any('mean' in s or 'std' in s or 'accuracy' in s or 'precision' in s or 'recall' in s or 'f1' in s for s in first_row_str)
                    
                    # 如果有字符串，说明第一行可能是标题行，需要特殊处理
                    if has_strings and df.shape[0] > 1:
                        print(f"  - 警告: sheet '{sheet_name}' 的第一行可能包含标题信息")
                        # 这种情况下，我们可能需要将第一行作为数据的一部分
                        # 但根据描述，文件只有2行，所以df.shape[0]==1
                        # 因此这个情况不太可能出现
                    
                    # 设置模型名为文件名
                    df.insert(0, 'model', model_name)
                    
                    # 设置索引为模型名
                    df.set_index('model', inplace=True)
                    
                    # 将模型数据添加到对应的sheet数据中
                    if sheet_name not in all_data:
                        all_data[sheet_name] = []
                    all_data[sheet_name].append(df)
                    
        except Exception as e:
            print(f"处理文件 {file_name} 时出错: {e}")
            import traceback
            traceback.print_exc()
    
    # 如果没有数据，返回
    if not all_data:
        print("未找到有效数据")
        print("可能的原因:")
        print("1. 文件结构不符合预期")
        print("2. 文件路径不正确")
        print("3. 文件格式不是标准的Excel格式")
        return
    
    # 创建新的工作簿
    wb = Workbook()
    wb.remove(wb.active)  # 移除默认的sheet
    
    # 处理每个sheet
    for sheet_name, dfs in all_data.items():
        print(f"正在处理指标: {sheet_name}")
        
        # 合并该sheet的所有数据
        if dfs:
            # 合并所有DataFrame
            combined_df = pd.concat(dfs, axis=0)
            
            # 按照模型名排序
            combined_df = combined_df.sort_index()
            
            # 创建sheet
            ws = wb.create_sheet(title=sheet_name)
            
            # 将DataFrame写入sheet
            for r_idx, row in enumerate(dataframe_to_rows(combined_df.reset_index(), index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            
            # 高亮显示最大值和第二大值
            highlight_max_values(ws, combined_df, sheet_name)
            
            print(f"  - 处理完成: {sheet_name}, 包含 {len(combined_df)} 个模型")
    
    # 保存工作簿
    wb.save(output_file)
    
    print(f"\n合并完成！结果已保存到: {output_file}")
    
    # 显示输出文件结构
    print("\n输出文件结构:")
    result_excel = openpyxl.load_workbook(output_file, read_only=True)
    for sheet_name in result_excel.sheetnames:
        ws = result_excel[sheet_name]
        print(f"  - {sheet_name}: {ws.max_row-1} 行数据")
    result_excel.close()


def main():
    # 设置输入文件夹和输出文件路径
    input_folder = input("请输入包含Excel文件的文件夹路径: ").strip()
    
    # 检查输入文件夹是否存在
    if not os.path.exists(input_folder):
        print(f"错误: 文件夹 {input_folder} 不存在")
        return
    
    # 自动生成输出文件名
    output_file = os.path.join('./', "result_excel.xlsx")
    
    print(f"输出文件将保存为: {output_file}")
    print("注意: 最大值将用黄色高亮，第二大值将用绿色高亮")
    
    # 确认是否继续
    response = input("是否继续? (y/n): ").strip().lower()
    if response != 'y':
        print("操作已取消")
        return
    
    # 执行合并
    merge_excel_files(input_folder, output_file)


if __name__ == "__main__":
    main()