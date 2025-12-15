import os
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

"""
合并目标检测任务的每一个指标
"""

def merge_excel_files(folder_path, output_file="merged_results.xlsx"):
    """
    合并指定文件夹中所有Excel文件的第二个工作簿
    
    参数:
    folder_path: Excel文件所在的文件夹路径
    output_file: 输出的合并文件名
    """
    
    # 获取文件夹中所有Excel文件
    excel_files = [f for f in os.listdir(folder_path) if f.endswith(('.xlsx', '.xls'))]
    
    if not excel_files:
        print("文件夹中没有找到Excel文件")
        return
    
    print(f"找到 {len(excel_files)} 个Excel文件")
    
    # 用于存储所有文件的数据
    all_data = {}
    
    # 读取第一个文件的第一列（模型和指标名称）
    first_file_path = os.path.join(folder_path, excel_files[0])
    first_df = pd.read_excel(first_file_path, sheet_name=1, header=None)
    model_column = first_df[0]
    
    # 处理每个Excel文件
    for file_name in excel_files:
        file_path = os.path.join(folder_path, file_name)
        print(f"正在处理: {file_name}")
        
        try:
            # 读取第二个工作簿
            df = pd.read_excel(file_path, sheet_name=1, header=None)
            
            # 获取文件名（不带扩展名）
            file_name_without_ext = os.path.splitext(file_name)[0]
            
            # 提取第二列数据
            if df.shape[1] > 1:
                # 第一行是文件名，其余是数据
                data_column = df[1].copy()
                all_data[file_name_without_ext] = data_column
            else:
                print(f"警告: {file_name} 没有第二列数据")
                
        except Exception as e:
            print(f"读取文件 {file_name} 时出错: {e}")
    
    # 创建合并后的DataFrame
    merged_df = pd.DataFrame()
    merged_df['Model'] = model_column
    
    # 将每个文件的数据作为新列添加到DataFrame中
    for file_name, data in all_data.items():
        merged_df[file_name] = data.values
    
    # 创建新的Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "Merged_Results"
    
    # 将DataFrame写入工作表
    for r_idx, row in enumerate(dataframe_to_rows(merged_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    # 获取指标行的索引（从第二行开始是指标）
    indicator_rows = []
    for i, value in enumerate(model_column.values, 1):
        if i > 1:  # 跳过第一行（表头"Model"）
            indicator_rows.append(i)
    
    # 获取数据列的范围（从第二列开始）
    data_columns_start = 2  # B列
    data_columns_end = len(all_data) + 1  # 根据文件数量确定
    
    # 对每个指标行应用格式
    for row_idx in indicator_rows:
        # 获取该行所有数据值
        row_values = []
        for col_idx in range(data_columns_start, data_columns_end + 1):
            cell_value = ws.cell(row=row_idx+1, column=col_idx).value  # +1因为Excel行从1开始
            if cell_value is not None:
                try:
                    row_values.append((col_idx, float(cell_value)))
                except:
                    row_values.append((col_idx, 0))  # 如果转换失败，设为0
        
        # 按值排序（从大到小）
        sorted_values = sorted(row_values, key=lambda x: x[1], reverse=True)
        
        # 为前三个最大值应用格式
        if len(sorted_values) > 0:
            # 最大值 - 加粗
            max_col_idx = sorted_values[0][0]
            max_cell = ws.cell(row=row_idx+1, column=max_col_idx)
            max_cell.font = Font(bold=True)
        
        if len(sorted_values) > 1:
            # 第二大值 - 下划线
            second_max_col_idx = sorted_values[1][0]
            second_max_cell = ws.cell(row=row_idx+1, column=second_max_col_idx)
            second_max_cell.font = Font(underline='single')
        
        if len(sorted_values) > 2:
            # 第三大值 - 斜体
            third_max_col_idx = sorted_values[2][0]
            third_max_cell = ws.cell(row=row_idx+1, column=third_max_col_idx)
            third_max_cell.font = Font(italic=True)
    
    # 保存文件
    wb.save(output_file)
    print(f"\n合并完成！结果已保存到: {output_file}")
    
    # 打印合并后的数据预览
    print("\n合并后的数据预览:")
    print(merged_df.to_string(index=False))

def main():
    # 设置文件夹路径
    folder_path = input("请输入包含Excel文件的文件夹路径: ").strip()
    
    # 验证文件夹是否存在
    if not os.path.exists(folder_path):
        print(f"错误: 文件夹 '{folder_path}' 不存在")
        return
    
    # 设置输出文件名
    output_file = input("请输入输出文件名（默认为 merged_results.xlsx）: ").strip()
    if not output_file:
        output_file = "merged_results.xlsx"
    
    # 确保输出文件有.xlsx扩展名
    if not output_file.endswith('.xlsx'):
        output_file += '.xlsx'
    
    # 执行合并
    merge_excel_files(folder_path, output_file)

if __name__ == "__main__":
    main()