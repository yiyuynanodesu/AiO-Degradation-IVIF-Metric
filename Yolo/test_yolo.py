from ultralytics import YOLO
import numpy as np
from PIL import Image
from tqdm import tqdm
import os
import torch
import warnings
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

def write_excel(excel_name='detection.xlsx', worksheet_name='VIF', column_index=0, data=None):
    try:
        workbook = load_workbook(excel_name)
    except FileNotFoundError:
        workbook = Workbook()

    worksheet = workbook.create_sheet(title=worksheet_name) if worksheet_name not in workbook.sheetnames else workbook[
        worksheet_name]

    column = get_column_letter(column_index + 1)
    for i, value in enumerate(data):
        cell = worksheet[column + str(i + 1)]
        cell.value = value

    workbook.save(excel_name)

image_path = './dataset/MSRS/detection/images'
Method = 'model'
model = YOLO("yolov8n.pt")

# metrics = model.val(data="yolo_xml.yaml",plots=False)
metrics = model.val(data="yolo_txt.yaml",plots=False)

P, R, mAP50, mAP5095 = metrics.results_dict['metrics/precision(B)'], metrics.results_dict['metrics/recall(B)'], metrics.results_dict['metrics/mAP50(B)'], metrics.results_dict['metrics/mAP50-95(B)']
mAP75 = metrics.box.map75

print(f'P : {P} , R: {R} , mAP50: {mAP50}, mAP75: {mAP75}, mAP5095: {mAP5095}')

metric_list = ['Model','P','R','mAP50','mAP75','mAP5095']
value_list = [Method, P, R, mAP50, mAP75, mAP5095]

write_excel(Method + '_detection.xlsx', 'Detection', 0, metric_list)
write_excel(Method + '_detection.xlsx', 'Detection', 1, value_list)

model.predict(image_path,save=True)