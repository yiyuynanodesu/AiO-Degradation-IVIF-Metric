# coding:utf-8
import torch
import os
import argparse
import time
import numpy as np

from core.model_fusion import Network3

os.environ['CUDA_VISIBLE_DEVICES'] = '0'
import torch.nn.functional as F
from torch.autograd import Variable
from torch.utils.data import DataLoader

from TaskFusion_dataset2 import Fusion_dataset

from tqdm import tqdm
from torch.autograd import Variable
from PIL import Image
import os, argparse, time, datetime, sys, shutil, stat
from torch.autograd import Variable
from torch.utils.data import DataLoader

# from model_fusion_seg_tzy4 import Network
from util.util import compute_results, visualize
from sklearn.metrics import confusion_matrix
from scipy.io import savemat

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

def write_excel(excel_name='metric.xlsx', worksheet_name='VIF', column_index=0, data=None):
    try:
        workbook = load_workbook(excel_name)
    except FileNotFoundError:
        workbook = Workbook()

    worksheet = workbook.create_sheet(title=worksheet_name) if worksheet_name not in workbook.sheetnames else workbook[
        worksheet_name]

    column = get_column_letter(column_index + 1)
    for i, value in enumerate(data):
        cell = worksheet[column + str(i + 1)]
        cell.value = value

    workbook.save(excel_name)


def val_segformer2(args, model, class_text):
    n_class = len(class_text)
    conf_total = np.zeros((n_class, n_class))

    class_dict = {}
    for i in range(len(class_text)):
        class_dict[i] = class_text[i]

    score_thres = 0.7
    ignore_idx = 255
    h=480
    w=640

    image_size=(h,w)
    n_min = 8 * 640 * 480 // 8

    device = torch.device("cuda:{}".format(args.gpu) if torch.cuda.is_available() else "cpu")
    if args.gpu >= 0:
        model.eval().to(device)

    test_dataset = Fusion_dataset(result_path=args.result_path, label_path= args.label_path)
    test_loader = DataLoader(
        dataset=test_dataset,
        batch_size=args.batch_size,
        shuffle=False,
        num_workers=args.num_workers,
        pin_memory=True,
        drop_last=False,
    )
    test_loader.n_iter = len(test_loader)
    print("Im working (^w^)")
    with torch.no_grad():
        for it, (images_result,label,name) in enumerate(test_loader):
            images_result = Variable(images_result)
            if args.gpu >= 0:
                images_result = images_result.to(device)

            logits,_, seg1, = model.forward(images_result)
            seg1 = F.interpolate(seg1, size=label.shape[1:], mode='bilinear', align_corners=False)

            label = label.cpu().numpy().squeeze().flatten()
            prediction = seg1.argmax(
                1).cpu().numpy().squeeze().flatten() 
            conf = confusion_matrix(y_true=label, y_pred=prediction, labels=[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14])
            conf_total += conf


        precision_per_class, recall_per_class, iou_per_class = compute_results(conf_total)

        print('\n############################Class Metric####################################')
        print("*precision_per_class: \n")
        for i in range(len(precision_per_class)):
            print(f'{class_dict[i]} : Precision: {precision_per_class[i]} Recall: {recall_per_class[i]} IOU: {iou_per_class[i]} ')
        print('\n################################Average####################################')
        print("\n* average values (np.mean(x)): \n ACC: %.6f, iou: %.6f" \
              % (precision_per_class.mean(), iou_per_class.mean()))
        print("* average values (np.mean(np.nan_to_num(x))): \n ACC: %.6f, iou: %.6f" \
              % (np.mean(np.nan_to_num(precision_per_class)), np.mean(np.nan_to_num(iou_per_class))))

        # 写入excel
        sheet_list = ['precision', 'recall', 'iou']
        value_dict = {
            'precision': precision_per_class,
            'recall': recall_per_class,
            'iou': iou_per_class
        }
        # 哪些指标需要计算mean
        need_mean = ['precision', 'iou']
        for sheet in sheet_list:
            model_list = ['', args.model_name]
            cls_value_list = []
            cls_value_list.append(model_list)
            value_list = value_dict[sheet]
                
            for idx in range(len(class_text)):
                # 忽略 bicycle
                if class_text[idx] == 'bicycle':
                    continue
                temp_list = [class_text[idx], value_list[idx]]
                cls_value_list.append(temp_list)
            if sheet in need_mean:
                temp_list = ['mean', value_list.mean()]
                cls_value_list.append(temp_list)
                
            for idx in range(len(cls_value_list)):
                write_excel(args.metric_save_name, sheet, idx, cls_value_list[idx])
        
        return iou_per_class.mean()

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Run SeAFusiuon with pytorch')
    parser.add_argument('--model_name', '-M', type=str, default='model')
    parser.add_argument('--batch_size', '-B', type=int, default=1)
    parser.add_argument('--gpu', '-G', type=int, default=0)
    parser.add_argument('--num_workers', '-J', type=int, default=8)
    parser.add_argument('--result_path', '-R', type=str, default='')
    parser.add_argument('--label_path', '-L', type=str, default='')
    parser.add_argument('--save_path', '-S', type=str, default='./')
    args = parser.parse_args()
    save_name = f'{args.model_name}_segmentation.xlsx'
    args.metric_save_name = os.path.join(args.save_path, save_name)
    class_text = ['background', 'road', 'sidewalk', 'building', 'lamp', 'sign', 'vegetation', 'sky', 'person', 'car', 'truck', 'bus', 'motocycle', 'bicycle', 'pole']
    seg_model_path = './pretrained/model-fusion_add_final2.pth'
    print('| testing %s on GPU #%d with pytorch' % (args.model_name, args.gpu))
    model_seg = Network3('mit_b3', len(class_text)).cuda()
    model_seg.load_state_dict(torch.load(seg_model_path))
    val_segformer2(args, model_seg, class_text)
