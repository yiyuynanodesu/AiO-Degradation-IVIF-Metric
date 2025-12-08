import numpy as np
from PIL import Image
from .Metric_torch import *
from tqdm import tqdm
import os
import torch
import warnings
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")


def write_excel(excel_name='metric.xlsx', worksheet_name='VIF', column_index=0, data=None):
    try:
        workbook = load_workbook(excel_name)
    except FileNotFoundError:
        workbook = Workbook()

    worksheet = workbook.create_sheet(title=worksheet_name) if worksheet_name not in workbook.sheetnames else workbook[
        worksheet_name]

    column = get_column_letter(column_index + 1)
    for i, value in enumerate(data):
        cell = worksheet[column + str(i + 1)]
        cell.value = value

    workbook.save(excel_name)


def evaluation_one(f_name):
    f_img = Image.open(f_name).convert('L')

    f_img_tensor = torch.tensor(np.array(f_img)).float().to(device)

    EI = EI_function(f_img_tensor)
    EN = EN_function(f_img_tensor)
    SF = SF_function(f_img_tensor)
    SD = SD_function(f_img_tensor)
    AG = AG_function(f_img_tensor)

    return EI, EN, SF, SD, AG

def eval_batch(output_path, save_dir, degard_list=['HazeRain','HazeLow','Rain','Haze','Exposure','Light'], model_name='Model', excel_filename='metric.xlsx'):
    Method_list = [
        model_name
    ]
    metric_save_name = os.path.join(save_dir, excel_filename)
    # Starting index for the method 'BDLFusion'
    start_index = Method_list.index(model_name)
    for i, Method in enumerate(Method_list[start_index:], start=start_index):
        EI_list = []
        EN_list = []
        SF_list = []
        AG_list = []
        SD_list = []
        filename_list = []
        for _ in range(len(degard_list)):
            EI_list.append([])
            EN_list.append([])
            SF_list.append([])
            AG_list.append([])
            SD_list.append([])
            filename_list.append([])
        
        
        output_fileList = os.listdir(output_path)
        eval_bar = tqdm(output_fileList)
        for _, output_filename in enumerate(eval_bar):
            f_name = os.path.join(output_path, output_filename)

            # 0--HazeRain 1--HazeLow 2--Rain 3--Haze 4--Exposure 5--Light
            degard_type = None

            if "Rain" in output_filename and "Haze" in output_filename:
                degard_type = 0
            elif "Low" in output_filename and "Haze" in output_filename:
                degard_type = 1
            elif "Rain" in output_filename:
                degard_type = 2
            elif "Haze" in output_filename:
                degard_type = 3
            elif "exposure" in output_filename:
                degard_type = 4
            elif "light" in output_filename:
                degard_type = 5

            if degard_type == None:
                print(f'!!! pay attention {f_name} can not classify!!!')
            print(f'{output_filename} {degard_type}')
            EI, EN, SF, SD, AG = evaluation_one(f_name)
            EI_list[degard_type].append(EI)
            EN_list[degard_type].append(EN)
            SF_list[degard_type].append(SF)
            AG_list[degard_type].append(AG)
            SD_list[degard_type].append(SD)
            filename_list[degard_type].append(output_filename)
            eval_bar.set_description("{} | {} | ".format(Method, output_filename, {degard_list[degard_type]}))

        for idx in range(len(degard_list)):
            EI_tensor = torch.tensor(EI_list[idx]).mean().item()
            EI_list[idx].append(EI_tensor)
            EN_tensor = torch.tensor(EN_list[idx]).mean().item()
            EN_list[idx].append(EN_tensor)
            SF_tensor = torch.tensor(SF_list[idx]).mean().item()
            SF_list[idx].append(SF_tensor)
            AG_tensor = torch.tensor(AG_list[idx]).mean().item()
            AG_list[idx].append(AG_tensor)
            SD_tensor = torch.tensor(SD_list[idx]).mean().item()
            SD_list[idx].append(SD_tensor)
            filename_list[idx].append('mean')

            filename_list[idx].insert(0, '')
            EI_list[idx].insert(0, 'EI')
            EN_list[idx].insert(0, 'EN')
            SF_list[idx].insert(0, 'SF')
            AG_list[idx].insert(0, 'AG')
            SD_list[idx].insert(0, 'SD')

        for idx in range(len(degard_list)):
            write_excel(metric_save_name, degard_list[idx], 0, filename_list[idx])
            write_excel(metric_save_name, degard_list[idx], 1, [x.item() if isinstance(x, torch.Tensor) else float(x) if isinstance(x, (int, float)) else x for x
                 in EI_list[idx]])
            write_excel(metric_save_name, degard_list[idx], 2, [x.item() if isinstance(x, torch.Tensor) else float(x) if isinstance(x, (int, float)) else x for x
                 in EN_list[idx]])
            write_excel(metric_save_name, degard_list[idx], 3, [x.item() if isinstance(x, torch.Tensor) else float(x) if isinstance(x, (int, float)) else x for x
                 in SF_list[idx]])
            write_excel(metric_save_name, degard_list[idx], 4, [x.item() if isinstance(x, torch.Tensor) else float(x) if isinstance(x, (int, float)) else x for x
                 in AG_list[idx]])
            write_excel(metric_save_name, degard_list[idx], 5, [x.item() if isinstance(x, torch.Tensor) else float(x) if isinstance(x, (int, float)) else x for x
                 in SD_list[idx]])
        print('Done')