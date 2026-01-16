import numpy as np
from PIL import Image
from .Metric_torch import *
from tqdm import tqdm
import os
import torch
import warnings
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from CLIPIQA.demo.clipiqa_single_image_demo import evaluate_image as CLIPIQA_eval
from TReSIQA.testing import evaluate_image as TReS_eval

warnings.filterwarnings("ignore")
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

def write_excel(excel_name='metric.xlsx', worksheet_name='VIF', column_index=0, data=None):
    try:
        workbook = load_workbook(excel_name)
    except FileNotFoundError:
        workbook = Workbook()

    worksheet = workbook.create_sheet(title=worksheet_name) if worksheet_name not in workbook.sheetnames else workbook[
        worksheet_name]

    column = get_column_letter(column_index + 1)
    for i, value in enumerate(data):
        cell = worksheet[column + str(i + 1)]
        cell.value = value

    workbook.save(excel_name)


def evaluation_one(f_name):
    f_img = Image.open(f_name).convert('L')

    f_img_tensor = torch.tensor(np.array(f_img)).float().to(device)

    EI = EI_function(f_img_tensor)
    EN = EN_function(f_img_tensor)
    SF = SF_function(f_img_tensor)
    SD = SD_function(f_img_tensor)
    AG = AG_function(f_img_tensor)
    CLIPIQA = CLIPIQA_eval(f_name)
    TReS = TReS_eval(f_name)

    return EI, EN, SF, SD, AG, CLIPIQA, TReS

def eval_batch(output_path, save_dir, model_name='Model', excel_filename='metric.xlsx'):
    Method = model_name
    metric_save_name = os.path.join(save_dir, excel_filename)
    
    EI_list = []
    EN_list = []
    SF_list = []
    AG_list = []
    SD_list = []
    CLIPIQA_list = []
    TReS_list = []
    
    filename_list = ['']
    output_fileList = os.listdir(output_path)
    eval_bar = tqdm(output_fileList)
    for _, output_filename in enumerate(eval_bar):
        f_name = os.path.join(output_path, output_filename)
        
        EI, EN, SF, SD, AG, CLIPIQA, TReS = evaluation_one(f_name)
        EI_list.append(EI)
        EN_list.append(EN)
        SF_list.append(SF)
        AG_list.append(AG)
        SD_list.append(SD)
        CLIPIQA_list.append(CLIPIQA)
        TReS_list.append(TReS)
        filename_list.append(output_filename)
        eval_bar.set_description("{} | {}".format(Method, output_filename))

    EI_tensor = torch.tensor(EI_list).mean().item()
    EI_list.append(EI_tensor)
    EN_tensor = torch.tensor(EN_list).mean().item()
    EN_list.append(EN_tensor)
    SF_tensor = torch.tensor(SF_list).mean().item()
    SF_list.append(SF_tensor)
    AG_tensor = torch.tensor(AG_list).mean().item()
    AG_list.append(AG_tensor)
    SD_tensor = torch.tensor(SD_list).mean().item()
    SD_list.append(SD_tensor)
    CLIPIQA_tensor = torch.tensor(CLIPIQA_list).mean().item()
    CLIPIQA_list.append(CLIPIQA_tensor)
    TReS_tensor = torch.tensor(TReS_list).mean().item()
    TReS_list.append(TReS_tensor)
    filename_list.append('mean')


    EI_list.insert(0, '{}'.format(Method))
    EN_list.insert(0, '{}'.format(Method))
    SF_list.insert(0, '{}'.format(Method))
    AG_list.insert(0, '{}'.format(Method))
    SD_list.insert(0, '{}'.format(Method))
    CLIPIQA_list.insert(0, '{}'.format(Method))
    TReS_list.insert(0, '{}'.format(Method))


    write_excel(metric_save_name, 'EI', 0, filename_list)
    write_excel(metric_save_name, 'EN', 0, filename_list)
    write_excel(metric_save_name, "SF", 0, filename_list)
    write_excel(metric_save_name, "AG", 0, filename_list)
    write_excel(metric_save_name, "SD", 0, filename_list)
    write_excel(metric_save_name, "CLIPIQA", 0, filename_list)
    write_excel(metric_save_name, "TReS", 0, filename_list)

    write_excel(metric_save_name, 'EI', 1,
                [x.item() if isinstance(x, torch.Tensor) else float(x) if isinstance(x, (int, float)) else x for x
                 in EI_list])
    write_excel(metric_save_name, 'EN', 1,
                [x.item() if isinstance(x, torch.Tensor) else float(x) if isinstance(x, (int, float)) else x for x
                 in EN_list])
    write_excel(metric_save_name, 'SF', 1,
                [x.item() if isinstance(x, torch.Tensor) else float(x) if isinstance(x, (int, float)) else x for x
                 in SF_list])
    write_excel(metric_save_name, 'AG', 1,
                [x.item() if isinstance(x, torch.Tensor) else float(x) if isinstance(x, (int, float)) else x for x
                 in AG_list])
    write_excel(metric_save_name, 'SD', 1,
                [x.item() if isinstance(x, torch.Tensor) else float(x) if isinstance(x, (int, float)) else x for x
                 in SD_list])
    write_excel(metric_save_name, 'CLIPIQA', 1,
                [x.item() if isinstance(x, torch.Tensor) else float(x) if isinstance(x, (int, float)) else x for x
                 in CLIPIQA_list])
    write_excel(metric_save_name, 'CLIPIQA', 1,
                [x.item() if isinstance(x, torch.Tensor) else float(x) if isinstance(x, (int, float)) else x for x
                 in CLIPIQA_list])
    write_excel(metric_save_name, 'TReS', 1,
                [x.item() if isinstance(x, torch.Tensor) else float(x) if isinstance(x, (int, float)) else x for x
                 in TReS_list])
    print('Done ╰(*°▽°*)╯')